"""
QR Ticket Email Sender — Resend version (Gmail-compliant)
----------------------------------------------------------
Reads your Excel file and sends each student a beautiful
HTML ticket email with their personal QR code embedded.

SETUP:
  1. uv add resend qrcode pillow openpyxl python-dotenv
     OR: pip install resend qrcode[pil] openpyxl pillow python-dotenv
  2. Fill in your .env file
  3. python main.py

Column layout in students_database.xlsx:
  A — Name
  B — Class
  C — ID
  D — Email
  E — Sent (1 = email sent, 0 or empty = not sent yet)

Main send run skips anyone already marked 1 in column E.
After a successful send, column E is updated to 1.

To resend to one student regardless of their sent status:
  python main.py resend 123
  python main.py resend "Jana Nováková"
  python main.py resend jana@example.com
"""

import io
import base64
import os
import sys
import time
import openpyxl
import qrcode
import resend
from dotenv import load_dotenv

load_dotenv()

# ================================================================
# CONFIG
# ================================================================

EXCEL_FILE         = "students_database.xlsx"
TEMPLATE_FILE      = "ticket_template.html"

COL_NAME           = 1   # A
COL_CLASS          = 2   # B
COL_ID             = 3   # C
COL_EMAIL          = 4   # D
COL_SENT           = 5   # E  — 1 = sent, 0/empty = not sent

HEADER_ROW         = 1

RESEND_API_KEY     = os.getenv("RESEND_API_KEY")
SENDER_EMAIL       = os.getenv("SENDER_EMAIL")
SENDER_NAME        = os.getenv("SENDER_NAME")
EVENT_NAME         = os.getenv("EVENT_NAME")
EVENT_DATE         = os.getenv("EVENT_DATE")
EVENT_LOCATION     = os.getenv("EVENT_LOCATION")
EVENT_TIME         = os.getenv("EVENT_TIME")

SEND_DELAY_SECONDS = 1.5

# "preview" = save QR images + print to console, no emails sent
# "real"    = actually send emails
MODE = "preview"

# ================================================================


def load_template():
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(
            f"Template file '{TEMPLATE_FILE}' not found. "
            "Make sure it's in the same folder as this script."
        )
    with open(TEMPLATE_FILE, "r", encoding="utf-8") as f:
        return f.read()


def make_qr_bytes(name, class_, id_):
    text = f"{name} | {class_} | {id_}"
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=3,
    )
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1a1a2e", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def make_plain_text(name, class_, id_):
    return f"""Ahoj {name},

Tvoj vstupný lístok na {EVENT_NAME} je pripravený.

Podujatie: {EVENT_NAME}
Dátum:     {EVENT_DATE}
Čas:       {EVENT_TIME}
Miesto:    {EVENT_LOCATION}

Tvoje údaje:
  Meno:   {name}
  Trieda: {class_}
  ID:     {id_}

QR kód je priložený k tomuto emailu ako obrázok. Ukáž ho pri vstupe.

---
Tento email bol odoslaný automaticky systémom lístkov pre {EVENT_NAME}.
Odosielateľ: {SENDER_EMAIL}
"""


def load_students(path, unsent_only=False):
    """
    Load students from Excel.
    Returns a list of dicts, each with a 'row_number' key for writing back.
    If unsent_only=True, skips anyone with COL_SENT == 1.
    Also ensures the header row has 'Sent' in column E.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Make sure column E has a header
    if ws.cell(row=HEADER_ROW, column=COL_SENT).value is None:
        ws.cell(row=HEADER_ROW, column=COL_SENT).value = "Sent"
        wb.save(path)

    students = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1):
        name   = row[COL_NAME  - 1].value
        class_ = row[COL_CLASS - 1].value
        id_    = row[COL_ID    - 1].value
        email  = row[COL_EMAIL - 1].value
        sent   = row[COL_SENT  - 1].value
        row_num = row[0].row

        if not name and not email:
            continue
        if not name:
            print(f"  ⚠  Skipping row {row_num} — no name")
            continue
        if not email:
            print(f"  ⚠  Skipping {name} — no email address")
            continue

        already_sent = (str(sent).strip() == "1") if sent is not None else False

        if unsent_only and already_sent:
            continue

        students.append({
            "name":         str(name).strip(),
            "class_":       str(class_).strip(),
            "id":           str(id_).strip(),
            "email":        str(email).strip(),
            "already_sent": already_sent,
            "row_number":   row_num,
        })

    wb.close()
    return students


def mark_sent(path, row_number):
    """Write 1 into column E for the given row number."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.cell(row=row_number, column=COL_SENT).value = 1
    wb.save(path)
    wb.close()


def _send_one_email(s, template, index=None, total=None):
    """
    Build and send a single ticket email.
    Returns True on success, False on failure.
    Does NOT update the Excel — caller decides whether to mark_sent.
    """
    qr_bytes = make_qr_bytes(s["name"], s["class_"], s["id"])
    qr_b64   = base64.b64encode(qr_bytes).decode()
    qr_cid   = f"qr_{s['id']}@tickets.rozlucka.me"

    html = template.format(
        name=s["name"],
        class_=s["class_"],
        id_=s["id"],
        qr_cid=qr_cid,
        EVENT_NAME=EVENT_NAME,
        EVENT_DATE=EVENT_DATE,
        EVENT_TIME=EVENT_TIME,
        EVENT_LOCATION=EVENT_LOCATION,
        SENDER_EMAIL=SENDER_EMAIL,
    )
    plain   = make_plain_text(s["name"], s["class_"], s["id"])
    subject = f"Tvoj lístok na {EVENT_NAME} – {s['name']}"
    prefix  = f"[{index}/{total}] " if index and total else ""

    try:
        resend.Emails.send({
            "from": f"{SENDER_NAME} <{SENDER_EMAIL}>",
            "to": [s["email"]],
            "subject": subject,
            "html": html,
            "text": plain,
            "attachments": [
                {
                    "filename": f"listok_{s['name'].replace(' ', '_')}.png",
                    "content": qr_b64,
                    "content_type": "image/png",
                    "content_id": qr_cid,
                }
            ],
        })
        print(f"  {prefix}✓ Sent to {s['name']} <{s['email']}>")
        return True

    except Exception as e:
        print(f"  {prefix}✗ Failed for {s['name']}: {e}")
        return False


def print_client_info(s):
    """Print a student's details in a consistent format."""
    status = "already sent" if s.get("already_sent") else "not sent yet"
    print(f"  Name:   {s['name']}")
    print(f"  Class:  {s['class_']}")
    print(f"  ID:     {s['id']}")
    print(f"  Email:  {s['email']}")
    print(f"  Status: {status}")
    print()


def resend_one(identifier: str):
    """
    Resend the ticket email to exactly one client.

    - Searches by ID, full name, or email (case-insensitive).
    - Always shows client info, even in preview mode.
    - In preview mode: shows info but does not send.
    - In real mode: asks for confirmation, then sends regardless of sent status.
    - Does NOT update the sent column (resend is manual, intentional).

    Usage:
        python main.py resend 123
        python main.py resend "Jana Nováková"
        python main.py resend jana@example.com
    """
    template = load_template()
    # Load all students (not unsent_only — resend ignores sent status)
    students = load_students(EXCEL_FILE, unsent_only=False)

    needle = identifier.strip().lower()
    matches = [
        s for s in students
        if needle in (s["id"].lower(),
                      s["name"].lower(),
                      s["email"].lower())
    ]

    if not matches:
        print(f"  ✗ No student found matching '{identifier}'")
        print("    Searched by: ID, name, and email (case-insensitive)")
        return

    if len(matches) > 1:
        print(f"  ⚠  Multiple matches for '{identifier}':\n")

    for m in matches:
        print_client_info(m)

    if MODE == "preview":
        print("  ⚠  MODE is set to 'preview' — email not sent.")
        print("     Change MODE to 'real' in the script to send for real.")
        return

    check_env()
    resend.api_key = RESEND_API_KEY

    label = "all of the above" if len(matches) > 1 else "this student"
    confirm = input(f"  Resend ticket to {label}? Type YES to confirm: ")
    if confirm.strip().upper() != "YES":
        print("  Cancelled.")
        return

    for s in matches:
        _send_one_email(s, template)


def preview_mode(students, template):
    os.makedirs("qr_preview", exist_ok=True)
    print("\n── PREVIEW MODE — no emails sent ──\n")
    for s in students:
        qr_bytes = make_qr_bytes(s["name"], s["class_"], s["id"])
        fname = f"qr_preview/qr_{s['name'].replace(' ', '_')}.png"
        with open(fname, "wb") as f:
            f.write(qr_bytes)

        qr_cid = f"qr_{s['id']}@tickets.rozlucka.me"
        html_preview = template.format(
            name=s["name"],
            class_=s["class_"],
            id_=s["id"],
            qr_cid=qr_cid,
            EVENT_NAME=EVENT_NAME,
            EVENT_DATE=EVENT_DATE,
            EVENT_TIME=EVENT_TIME,
            EVENT_LOCATION=EVENT_LOCATION,
            SENDER_EMAIL=SENDER_EMAIL,
        )
        html_fname = f"qr_preview/preview_{s['name'].replace(' ', '_')}.html"
        with open(html_fname, "w", encoding="utf-8") as f:
            f.write(html_preview)

        print(f"TO:      {s['email']}")
        print(f"SUBJECT: Tvoj lístok na {EVENT_NAME} – {s['name']}")
        print(f"QR:      {fname}")
        print(f"HTML:    {html_fname}  ← open in browser to preview")
        print("─" * 50)
    print(f"\n✓ {len(students)} tickets previewed. Files saved to qr_preview/")
    print('  Set MODE = "real" to send for real.\n')


def send_mode(students, template):
    resend.api_key = RESEND_API_KEY

    sent_count = 0
    failed = 0

    for i, s in enumerate(students, 1):
        ok = _send_one_email(s, template, index=i, total=len(students))
        if ok:
            mark_sent(EXCEL_FILE, s["row_number"])
            sent_count += 1
        else:
            failed += 1

        if i < len(students):
            time.sleep(SEND_DELAY_SECONDS)

    print(f"\n── Done: {sent_count} sent, {failed} failed ──\n")


def check_env():
    missing = [k for k in [
        "RESEND_API_KEY", "SENDER_EMAIL", "SENDER_NAME",
        "EVENT_NAME", "EVENT_DATE", "EVENT_LOCATION", "EVENT_TIME"
    ] if not os.getenv(k)]
    if missing:
        print(f"  ✗ Missing .env variables: {', '.join(missing)}")
        print("    Fill them in your .env file and try again.")
        exit(1)


def main():
    # Handle: python main.py resend <identifier>
    if len(sys.argv) >= 3 and sys.argv[1].lower() == "resend":
        identifier = " ".join(sys.argv[2:])
        resend_one(identifier)
        return

    check_env()

    print(f"Loading template from {TEMPLATE_FILE} …")
    template = load_template()

    print(f"Loading {EXCEL_FILE} …")

    if MODE == "preview":
        # Preview shows all students regardless of sent status
        students = load_students(EXCEL_FILE, unsent_only=False)
        print(f"  Found {len(students)} students.\n")
        if not students:
            print("No students found. Check your column settings in CONFIG.")
            return
        preview_mode(students, template)

    elif MODE == "real":
        # Only load students who haven't been sent to yet
        students = load_students(EXCEL_FILE, unsent_only=True)
        print(f"  Found {len(students)} unsent students.\n")
        if not students:
            print("  ✓ All students have already been sent an email. Nothing to do.")
            return
        confirm = input(f"Send emails to {len(students)} unsent students? Type YES to confirm: ")
        if confirm.strip().upper() == "YES":
            send_mode(students, template)
        else:
            print("Cancelled.")

    else:
        print(f"Unknown MODE '{MODE}'. Use 'preview' or 'real'.")


if __name__ == "__main__":
    main()
