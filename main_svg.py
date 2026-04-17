"""
QR Ticket Email Sender — SVG Template Version
----------------------------------------------
Reads your Excel file and sends each student a beautiful
ticket email with their personal QR code embedded in an SVG ticket.

SETUP:
  1. pip install resend qrcode[pil] openpyxl pillow python-dotenv playwright
     playwright install chromium
  2. Fill in your .env file
  3. Make sure your SVG template has these placeholders:
     {NAME_PLACEHOLDER} — will be replaced with student name
     {CLASS_PLACEHOLDER} — will be replaced with class
  4. python main_svg.py

Column layout in students_database.xlsx:
  A — Name
  B — Class
  C — ID
  D — Email
  E — Sent (1 = email sent, 0 or empty = not sent yet)

Main send run skips anyone already marked 1 in column E.
After a successful send, column E is updated to 1.

To resend to one student regardless of their sent status:
  python main_svg.py resend 123
  python main_svg.py resend "Jana Nováková"
  python main_svg.py resend jana@example.com
"""

import io
import base64
import os
import sys
import time
import tempfile
import openpyxl
import qrcode
import resend
from dotenv import load_dotenv
from PIL import Image
from playwright.sync_api import sync_playwright

load_dotenv()

# ================================================================
# CONFIG
# ================================================================

EXCEL_FILE         = "students_database.xlsx"
SVG_TEMPLATE_FILE  = "ticket_template.svg"

COL_NAME           = 1   # A
COL_CLASS          = 2   # B
COL_ID             = 3   # C
COL_EMAIL          = 4   # D
COL_SENT           = 5   # E  — 1 = sent, 0/empty = not sent

HEADER_ROW         = 1

RESEND_API_KEY     = os.getenv("RESEND_API_KEY")
SENDER_EMAIL       = os.getenv("SENDER_EMAIL")
SENDER_NAME        = os.getenv("SENDER_NAME")
EVENT_NAME         = os.getenv("EVENT_NAME")
EVENT_DATE         = os.getenv("EVENT_DATE")
EVENT_LOCATION     = os.getenv("EVENT_LOCATION")
EVENT_TIME         = os.getenv("EVENT_TIME")

SEND_DELAY_SECONDS = 1.5

# How long the event lasts — used for the calendar invite end time
EVENT_DURATION_MINUTES = int(os.getenv("EVENT_DURATION_MINUTES", "120"))

# "preview" = save ticket files + print to console, no emails sent
# "real"    = actually send emails
MODE = "preview"

# Shift ALL text elements up (negative) or down (positive) in the PDF output.
# SVGs look fine but Playwright's print renderer can push text down slightly.
# Run:  python main_svg.py adjustpdf
# Open ticket_preview/_adjust_test.pdf, tweak this value, repeat until perfect.
# Typical fix range: -5 to -20 (usually negative = move up).
TEXT_Y_OFFSET = -17   # pixels

# ================================================================


def load_svg_template():
    if not os.path.exists(SVG_TEMPLATE_FILE):
        raise FileNotFoundError(
            f"SVG template file '{SVG_TEMPLATE_FILE}' not found. "
            "Make sure it's in the same folder as this script."
        )
    with open(SVG_TEMPLATE_FILE, "r", encoding="utf-8") as f:
        return f.read()


def make_qr_bytes(name, class_, id_, size_px=400):
    text = f"{name} | {class_} | {id_}"
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1a1a2e", back_color="white")
    img = img.resize((size_px, size_px), Image.Resampling.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def qr_to_base64(qr_bytes):
    b64 = base64.b64encode(qr_bytes).decode()
    return f"data:image/png;base64,{b64}"




def create_svg_ticket(svg_template, name, class_, id_):
    qr_bytes = make_qr_bytes(name, class_, id_, size_px=350)
    qr_data_uri = qr_to_base64(qr_bytes)

    svg = svg_template
    svg = svg.replace("{NAME_PLACEHOLDER}", name)
    svg = svg.replace("{CLASS_PLACEHOLDER}", class_)
    svg = svg.replace("{QR_CODE_DATA}", qr_data_uri)
    return svg


def _apply_text_y_offset(svg_content, delta):
    """
    Shift every <text> and <tspan> y attribute in the SVG source by `delta`
    pixels.  Works directly on the coordinate values so it's immune to any
    existing transform= attributes Figma may have placed on those elements.
    Positive delta = move down, negative = move up.
    """
    if delta == 0:
        return svg_content

    import re as _re

    def shift_y_attr(m):
        # m.group(1) = everything before the y value list
        # m.group(2) = the space-separated list of y values
        prefix = m.group(1)
        values = m.group(2).strip().split()
        try:
            shifted = ' '.join(f'{float(v) + delta:.4f}' for v in values)
        except ValueError:
            return m.group(0)   # not numeric — leave untouched
        return f'{prefix}"{shifted}"'

    # Match y="..." on <text> and <tspan> tags only
    # Pattern captures the tag-and-attribute prefix so we don't touch
    # y attributes on unrelated elements (e.g. <rect y="...">)
    svg_content = _re.sub(
        r'(<(?:text|tspan)[^>]*?\by=)"([^"]*)"',
        shift_y_attr,
        svg_content,
    )
    # Also handle single-quoted y='...'
    def shift_y_attr_sq(m):
        prefix = m.group(1)
        values = m.group(2).strip().split()
        try:
            shifted = ' '.join(f'{float(v) + delta:.4f}' for v in values)
        except ValueError:
            return m.group(0)
        return f"{prefix}'{shifted}'"

    svg_content = _re.sub(
        r"(<(?:text|tspan)[^>]*?\by=)'([^']*)'",
        shift_y_attr_sq,
        svg_content,
    )
    return svg_content


def svg_to_pdf_bytes(svg_content):
    """
    Convert SVG string to PDF bytes using a headless Chromium browser
    via Playwright.  Renders with full browser fidelity — gradients,
    masks, blend modes, complex paths and all Figma SVG features are
    supported correctly.

    TEXT_Y_OFFSET is applied by directly nudging y attributes in the SVG
    source before rendering, so it never interferes with Figma's existing
    transform attributes.

    Requires:
        pip install playwright
        playwright install chromium
    """
    # Apply y offset directly in SVG source — safe, no CSS transform conflicts
    if TEXT_Y_OFFSET != 0:
        svg_content = _apply_text_y_offset(svg_content, TEXT_Y_OFFSET)

    # Parse width/height from the SVG so we can size the HTML page to match
    import re as _re
    w_px, h_px = 1200, 600  # safe fallback

    vb_match = _re.search(r'viewBox=["\'][\d.]+\s+[\d.]+\s+([\d.]+)\s+([\d.]+)["\']', svg_content)
    if vb_match:
        w_px = float(vb_match.group(1))
        h_px = float(vb_match.group(2))
    else:
        w_match = _re.search(r'<svg[^>]+\bwidth=["\']([0-9.]+)', svg_content)
        h_match = _re.search(r'<svg[^>]+\bheight=["\']([0-9.]+)', svg_content)
        if w_match:
            w_px = float(w_match.group(1))
        if h_match:
            h_px = float(h_match.group(1))

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  *, html, body {{
    margin: 0;
    padding: 0;
    border: 0;
    box-sizing: border-box;
  }}
  body {{
    width: {w_px}px;
    height: {h_px}px;
    overflow: hidden;
    background: transparent;
  }}
  svg {{
    display: block;
    width: {w_px}px;
    height: {h_px}px;
  }}
</style>
</head>
<body>
{svg_content}
</body>
</html>"""

    with tempfile.NamedTemporaryFile(
        suffix=".html", delete=False, mode="w", encoding="utf-8"
    ) as tmp:
        tmp.write(html)
        tmp_path = tmp.name

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()

            page.goto(f"file:///{tmp_path.replace(os.sep, '/')}")
            page.wait_for_timeout(300)

            page.set_viewport_size({"width": int(w_px), "height": int(h_px)})

            pdf_bytes = page.pdf(
                width=f"{w_px}px",
                height=f"{h_px}px",
                print_background=True,
                margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
            )

            browser.close()
    finally:
        os.unlink(tmp_path)

    return pdf_bytes


def make_plain_text(name, class_, id_):
    return f"""Ahoj {name},

Tvoj vstupný lístok na {EVENT_NAME} je pripravený.

Podujatie: {EVENT_NAME}
Dátum:     {EVENT_DATE}
Čas:       {EVENT_TIME}
Miesto:    {EVENT_LOCATION}

Tvoje údaje:
  Meno:   {name}
  Trieda: {class_}
  ID:     {id_}

Lístok je priložený k tomuto emailu ako PDF.
Pozvánka do kalendára je tiež priložená — stačí ju otvoriť. Ukáž lístok pri vstupe.

---
Tento email bol odoslaný automaticky systémom lístkov pre {EVENT_NAME}.
Odosielateľ: {SENDER_EMAIL}
"""


def make_ics(name, id_):
    """
    Build an iCalendar (.ics) file as a string.

    EVENT_DATE must be in the format set in your .env, e.g. "15.6.2026" or "2026-06-15".
    EVENT_TIME must be "HH:MM", e.g. "18:00".

    The ics is personalised per student (UID includes their ID) so duplicate
    detection works correctly if they import it more than once.
    """
    import re as _re
    from datetime import datetime, timedelta

    # Parse date — accept DD.MM.YYYY or YYYY-MM-DD
    date_str = (EVENT_DATE or "").strip()
    time_str = (EVENT_TIME or "00:00").strip()

    # Normalise Slovak long format: "Piatok, 30. mája 2026" → "30. mája 2026"
    # Strip optional leading weekday + comma
    date_str = _re.sub(r'^[^,]+,\s*', '', date_str).strip()

    # Replace Slovak month names with numbers
    SK_MONTHS = {
        "januára": "1", "februára": "2", "marca": "3", "apríla": "4",
        "mája": "5", "júna": "6", "júla": "7", "augusta": "8",
        "septembra": "9", "októbra": "10", "novembra": "11", "decembra": "12",
        # nominative forms just in case
        "január": "1", "február": "2", "marec": "3", "apríl": "4",
        "máj": "5", "jún": "6", "júl": "7", "august": "8",
        "september": "9", "október": "10", "november": "11", "december": "12",
    }
    for word, num in SK_MONTHS.items():
        date_str = _re.sub(rf'\b{word}\b', num, date_str, flags=_re.IGNORECASE)

    # Remove stray dots from "30. 5. 2026" → "30 5 2026", normalise separators
    date_str = _re.sub(r'\.', ' ', date_str)
    date_str = _re.sub(r'\s+', ' ', date_str).strip()

    dt = None
    for fmt in ("%d %m %Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(f"{date_str} {time_str}", f"{fmt} %H:%M")
            break
        except ValueError:
            continue

    if dt is None:
        # Fallback: can't parse — return empty string, skip the attachment
        print(f"  ⚠  Could not parse EVENT_DATE='{EVENT_DATE}' for calendar invite. Skipping ICS.")
        return None

    dt_end = dt + timedelta(minutes=EVENT_DURATION_MINUTES)

    fmt_ics = "%Y%m%dT%H%M%S"
    uid     = f"{id_}-{dt.strftime('%Y%m%d')}-{EVENT_NAME.replace(' ', '')}@ticket"

    # Fold long lines at 75 chars as required by RFC 5545
    def fold(line):
        if len(line.encode()) <= 75:
            return line
        chunks, out = [], []
        enc = line.encode()
        while enc:
            chunks.append(enc[:75].decode(errors='replace'))
            enc = enc[75:]
        return ('\r\n ').join(chunks)

    location = (EVENT_LOCATION or "").replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;")
    summary  = (EVENT_NAME or "").replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;")
    desc     = f"{name} | ID: {id_}".replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;")

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//QR Ticket Sender//SK",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "BEGIN:VEVENT",
        fold(f"UID:{uid}"),
        f"DTSTAMP:{datetime.utcnow().strftime(fmt_ics)}Z",
        f"DTSTART:{dt.strftime(fmt_ics)}",
        f"DTEND:{dt_end.strftime(fmt_ics)}",
        fold(f"SUMMARY:{summary}"),
        fold(f"LOCATION:{location}"),
        fold(f"DESCRIPTION:{desc}"),
        "STATUS:CONFIRMED",
        "END:VEVENT",
        "END:VCALENDAR",
    ]
    return "\r\n".join(lines) + "\r\n"


def load_students(path, unsent_only=False):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    if ws.cell(row=HEADER_ROW, column=COL_SENT).value is None:
        ws.cell(row=HEADER_ROW, column=COL_SENT).value = "Sent"
        wb.save(path)

    students = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1):
        name   = row[COL_NAME  - 1].value
        class_ = row[COL_CLASS - 1].value
        id_    = row[COL_ID    - 1].value
        email  = row[COL_EMAIL - 1].value
        sent   = row[COL_SENT  - 1].value
        row_num = row[0].row

        if not name and not email:
            continue
        if not name:
            print(f"  ⚠  Skipping row {row_num} — no name")
            continue
        if not email:
            print(f"  ⚠  Skipping {name} — no email address")
            continue

        already_sent = (str(sent).strip() == "1") if sent is not None else False

        if unsent_only and already_sent:
            continue

        students.append({
            "name":         str(name).strip(),
            "class_":       str(class_).strip(),
            "id":           str(id_).strip(),
            "email":        str(email).strip(),
            "already_sent": already_sent,
            "row_number":   row_num,
        })

    wb.close()
    return students


def mark_sent(path, row_number):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.cell(row=row_number, column=COL_SENT).value = 1
    wb.save(path)
    wb.close()


def _send_one_email(s, svg_template, index=None, total=None):
    svg_ticket = create_svg_ticket(svg_template, s["name"], s["class_"], s["id"])
    pdf_bytes  = svg_to_pdf_bytes(svg_ticket)
    pdf_b64    = base64.b64encode(pdf_bytes).decode()

    ics_content = make_ics(s["name"], s["id"])
    ics_b64     = base64.b64encode(ics_content.encode("utf-8")).decode() if ics_content else None

    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; text-align: center; padding: 40px; background: #f0f0f5;">
        <div style="max-width: 600px; margin: 0 auto;">
            <h1 style="color: #1a1a2e;">Ahoj {s["name"]}! 👋</h1>
            <p style="font-size: 18px; color: #555;">Tvoj vstupný lístok na <strong>{EVENT_NAME}</strong> je pripravený.</p>
            <div style="background: white; padding: 30px; border-radius: 12px; margin: 30px 0;">
                <p style="color: #999; font-size: 14px; margin: 10px 0;">📅 <strong>{EVENT_DATE}</strong> o <strong>{EVENT_TIME}</strong></p>
                <p style="color: #999; font-size: 14px; margin: 10px 0;">📍 <strong>{EVENT_LOCATION}</strong></p>
            </div>
            <p style="color: #666; font-size: 14px;">Tvoj osobný lístok je priložený k tomuto emailu ako PDF.</p>
            <p style="color: #666; font-size: 14px;">Ulož si ho a ukáž pri vstupe.</p>
            {"<p style='color: #666; font-size: 14px;'>📆 Priložená je aj <strong>pozvánka do kalendára</strong> — stačí ju otvoriť a udalosť sa pridá automaticky.</p>" if ics_b64 else ""}
            <p style="color: #aaa; font-size: 12px; margin-top: 40px;">
                Tešíme sa na teba! 🎉<br>
                <span style="color: #ccc; font-size: 11px;">ID: {s["id"]}</span>
            </p>
        </div>
    </body>
    </html>
    """

    plain   = make_plain_text(s["name"], s["class_"], s["id"])
    subject = f"Tvoj lístok na {EVENT_NAME} – {s['name']}"
    prefix  = f"[{index}/{total}] " if index and total else ""

    attachments = [
        {
            "filename": f"listok_{s['name'].replace(' ', '_')}_{s['id']}.pdf",
            "content":  pdf_b64,
            "content_type": "application/pdf",
        }
    ]
    if ics_b64:
        attachments.append({
            "filename": f"{EVENT_NAME.replace(' ', '_')}.ics",
            "content":  ics_b64,
            "content_type": "text/calendar",
        })

    try:
        resend.Emails.send({
            "from":        f"{SENDER_NAME} <{SENDER_EMAIL}>",
            "to":          [s["email"]],
            "subject":     subject,
            "html":        html_body,
            "text":        plain,
            "attachments": attachments,
        })
        print(f"  {prefix}✓ Sent to {s['name']} <{s['email']}>")
        return True

    except Exception as e:
        print(f"  {prefix}✗ Failed for {s['name']}: {e}")
        return False


def print_client_info(s):
    status = "already sent" if s.get("already_sent") else "not sent yet"
    print(f"  Name:   {s['name']}")
    print(f"  Class:  {s['class_']}")
    print(f"  ID:     {s['id']}")
    print(f"  Email:  {s['email']}")
    print(f"  Status: {status}")
    print()


def resend_one(identifier: str):
    svg_template = load_svg_template()
    students = load_students(EXCEL_FILE, unsent_only=False)

    needle  = identifier.strip().lower()
    matches = [
        s for s in students
        if needle in (s["id"].lower(), s["name"].lower(), s["email"].lower())
    ]

    if not matches:
        print(f"  ✗ No student found matching '{identifier}'")
        print("    Searched by: ID, name, and email (case-insensitive)")
        return

    if len(matches) > 1:
        print(f"  ⚠  Multiple matches for '{identifier}':\n")

    for m in matches:
        print_client_info(m)

    if MODE == "preview":
        print("  ⚠  MODE is set to 'preview' — email not sent.")
        print("     Change MODE to 'real' in the script to send for real.")
        return

    check_env()
    resend.api_key = RESEND_API_KEY

    label   = "all of the above" if len(matches) > 1 else "this student"
    confirm = input(f"  Resend ticket to {label}? Type YES to confirm: ")
    if confirm.strip().upper() != "YES":
        print("  Cancelled.")
        return

    for s in matches:
        _send_one_email(s, svg_template)


def preview_mode(students, svg_template):
    os.makedirs("ticket_preview", exist_ok=True)
    print("\n── PREVIEW MODE — no emails sent ──\n")

    for s in students:
        svg_ticket = create_svg_ticket(svg_template, s["name"], s["class_"], s["id"])

        safe_name = s["name"].replace(" ", "_")

        svg_fname = f"ticket_preview/ticket_{safe_name}.svg"
        with open(svg_fname, "w", encoding="utf-8") as f:
            f.write(svg_ticket)

        pdf_bytes = svg_to_pdf_bytes(svg_ticket)
        pdf_fname = f"ticket_preview/ticket_{safe_name}.pdf"
        with open(pdf_fname, "wb") as f:
            f.write(pdf_bytes)

        ics_content = make_ics(s["name"], s["id"])
        ics_fname = None
        if ics_content:
            ics_fname = f"ticket_preview/ticket_{safe_name}.ics"
            with open(ics_fname, "w", encoding="utf-8") as f:
                f.write(ics_content)

        print(f"TO:      {s['email']}")
        print(f"SUBJECT: Tvoj lístok na {EVENT_NAME} – {s['name']}")
        print(f"SVG:     {svg_fname}")
        print(f"PDF:     {pdf_fname}")
        if ics_fname:
            print(f"ICS:     {ics_fname}")
        print("─" * 50)

    print(f"\n✓ {len(students)} tickets previewed. Files saved to ticket_preview/")
    print('  Set MODE = "real" to send for real.\n')


def send_mode(students, svg_template):
    resend.api_key = RESEND_API_KEY
    sent_count = 0
    failed = 0

    for i, s in enumerate(students, 1):
        ok = _send_one_email(s, svg_template, index=i, total=len(students))
        if ok:
            mark_sent(EXCEL_FILE, s["row_number"])
            sent_count += 1
        else:
            failed += 1
        if i < len(students):
            time.sleep(SEND_DELAY_SECONDS)

    print(f"\n── Done: {sent_count} sent, {failed} failed ──\n")


def check_env():
    missing = [k for k in [
        "RESEND_API_KEY", "SENDER_EMAIL", "SENDER_NAME",
        "EVENT_NAME", "EVENT_DATE", "EVENT_LOCATION", "EVENT_TIME"
    ] if not os.getenv(k)]
    if missing:
        print(f"  ✗ Missing .env variables: {', '.join(missing)}")
        print("    Fill them in your .env file and try again.")
        exit(1)


def adjustpdf_mode():
    """
    Generate a single test PDF from the first student in the spreadsheet
    using the current TEXT_Y_OFFSET value.  Use this to dial in the offset
    before running the full batch.

    Usage:  python main_svg.py adjustpdf
    Output: ticket_preview/_adjust_test.pdf
    """
    print(f"\n── PDF ADJUSTMENT MODE ──")
    print(f"  Current TEXT_Y_OFFSET = {TEXT_Y_OFFSET} px")
    print(f"  Loading first student from {EXCEL_FILE} …\n")

    students = load_students(EXCEL_FILE, unsent_only=False)
    if not students:
        print("  ✗ No students found in the spreadsheet.")
        return

    s = students[0]
    print(f"  Using: {s['name']} ({s['class_']}, ID {s['id']})")

    svg_template = load_svg_template()
    svg_ticket   = create_svg_ticket(svg_template, s["name"], s["class_"], s["id"])

    os.makedirs("ticket_preview", exist_ok=True)
    out_path = "ticket_preview/_adjust_test.pdf"

    print(f"  Rendering PDF …")
    pdf_bytes = svg_to_pdf_bytes(svg_ticket)
    with open(out_path, "wb") as f:
        f.write(pdf_bytes)

    print(f"\n  ✓ Saved → {out_path}")
    print(f"\n  Open it, check the text position, then:")
    print(f"  • Text too LOW  → set TEXT_Y_OFFSET to a more negative number (e.g. -10, -15)")
    print(f"  • Text too HIGH → set TEXT_Y_OFFSET to a more positive number (e.g. +5)")
    print(f"  • Looks good    → leave TEXT_Y_OFFSET = {TEXT_Y_OFFSET} and run normally\n")


def main():
    if len(sys.argv) >= 2 and sys.argv[1].lower() == "adjustpdf":
        adjustpdf_mode()
        return

    if len(sys.argv) >= 3 and sys.argv[1].lower() == "resend":
        identifier = " ".join(sys.argv[2:])
        resend_one(identifier)
        return

    check_env()

    print(f"Loading SVG template from {SVG_TEMPLATE_FILE} …")
    svg_template = load_svg_template()

    print(f"Loading {EXCEL_FILE} …")

    if MODE == "preview":
        students = load_students(EXCEL_FILE, unsent_only=False)
        print(f"  Found {len(students)} students.\n")
        if not students:
            print("No students found. Check your column settings in CONFIG.")
            return
        preview_mode(students, svg_template)

    elif MODE == "real":
        students = load_students(EXCEL_FILE, unsent_only=True)
        print(f"  Found {len(students)} unsent students.\n")
        if not students:
            print("  ✓ All students have already been sent an email. Nothing to do.")
            return
        confirm = input(f"Send emails to {len(students)} unsent students? Type YES to confirm: ")
        if confirm.strip().upper() == "YES":
            send_mode(students, svg_template)
        else:
            print("Cancelled.")

    else:
        print(f"Unknown MODE '{MODE}'. Use 'preview' or 'real'.")


if __name__ == "__main__":
    main()