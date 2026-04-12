"""
Unique ID Generator for Excel
------------------------------
Fills empty cells in the ID column with unique random 3-digit numbers (100-999).
Already existing IDs are respected and never overwritten or duplicated.

Usage:
    python generate_ids.py students.xlsx
"""

import sys
import random
import openpyxl

EXCEL_FILE = "students_database.xlsx"   # change if needed, or pass as argument
COL_ID     = 3                 # column C
HEADER_ROW = 1

def generate_ids(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Collect all IDs that already exist
    used = set()
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        val = row[COL_ID - 1]
        if val is not None:
            used.add(int(val))

    # Build a pool of all available 3-digit numbers
    available = list(set(range(100, 1000)) - used)
    random.shuffle(available)

    filled = 0
    for row in ws.iter_rows(min_row=HEADER_ROW + 1):
        cell = row[COL_ID - 1]
        if cell.value is None or str(cell.value).strip() == "":
            if not available:
                print("  ✗ Ran out of unique 3-digit IDs (max 900 students)")
                break
            new_id = available.pop()
            cell.value = new_id
            used.add(new_id)
            filled += 1
            print(f"  Row {cell.row}: assigned ID {new_id}")

    wb.save(path)
    print(f"\n✓ Done — {filled} new IDs assigned, file saved.")

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_FILE
    generate_ids(path)
