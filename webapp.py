"""
Ticket Verifier — Flask backend
--------------------------------
Reads students_database.xlsx and exposes a simple API for the
mobile-friendly QR scanner web UI.

Column layout (same as Ticket-Sender):
  A — Name
  B — Class
  C — ID
  D — Email
  E — Sent  (1 = ticket was emailed)
  F — Arrived (1 = scanned at entrance, added by this app)

Run:
  pip install flask flask-cors openpyxl
  python app.py

Then open  http://<your-local-ip>:5000  on any device on the same WiFi.
"""

import os
import openpyxl
from flask import Flask, jsonify, request, render_template, send_from_directory
from flask_cors import CORS
import socket
import qrcode
socket.setdefaulttimeout(5)

app = Flask(__name__)
CORS(app)

# ── Config ──────────────────────────────────────────────────────────────────
EXCEL_FILE = os.environ.get("EXCEL_FILE", "students_database.xlsx")

COL_NAME    = 1   # A
COL_CLASS   = 2   # B
COL_ID      = 3   # C
COL_EMAIL   = 4   # D
COL_SENT    = 5   # E
COL_ARRIVED = 6   # F  ← new column managed by this app

HEADER_ROW  = 1
# ────────────────────────────────────────────────────────────────────────────


def _ensure_arrived_header():
    """Add 'Arrived' header in column F if it doesn't exist yet."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    if ws.cell(row=HEADER_ROW, column=COL_ARRIVED).value is None:
        ws.cell(row=HEADER_ROW, column=COL_ARRIVED).value = "Arrived"
        wb.save(EXCEL_FILE)
    wb.close()


def _load_all():
    """Return list of student dicts with row_number."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    students = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1):
        name    = row[COL_NAME    - 1].value
        class_  = row[COL_CLASS   - 1].value
        id_     = row[COL_ID      - 1].value
        email   = row[COL_EMAIL   - 1].value
        sent    = row[COL_SENT    - 1].value
        arrived = row[COL_ARRIVED - 1].value
        row_num = row[0].row

        if not name and not id_:
            continue

        students.append({
            "name":     str(name  or "").strip(),
            "class_":   str(class_ or "").strip(),
            "id":       str(id_   or "").strip(),
            "email":    str(email or "").strip(),
            "sent":     str(sent).strip() == "1" if sent is not None else False,
            "arrived":  str(arrived).strip() == "1" if arrived is not None else False,
            "row":      row_num,
        })
    wb.close()
    return students


def _mark_arrived(row_number: int):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.cell(row=row_number, column=COL_ARRIVED).value = 1
    wb.save(EXCEL_FILE)
    wb.close()


def _unmark_arrived(row_number: int):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.cell(row=row_number, column=COL_ARRIVED).value = 0
    wb.save(EXCEL_FILE)
    wb.close()


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/verify", methods=["POST"])
def verify():
    """
    Body: { "qr": "<raw QR string>" }
    QR format:  "Name | Class | ID"
    Returns ticket info + DB match status.
    """
    if not os.path.exists(EXCEL_FILE):
        return jsonify({"error": f"Database file '{EXCEL_FILE}' not found on server."}), 500

    data = request.get_json(force=True)
    qr_text = (data.get("qr") or "").strip()

    if not qr_text:
        return jsonify({"error": "Empty QR data"}), 400

    # Parse QR: "Name | Class | ID"
    parts = [p.strip() for p in qr_text.split("|")]
    if len(parts) != 3:
        return jsonify({
            "valid": False,
            "reason": "unrecognised_format",
            "raw": qr_text,
        })

    qr_name, qr_class, qr_id = parts

    _ensure_arrived_header()
    students = _load_all()

    match = next((s for s in students if s["id"] == qr_id), None)

    if match is None:
        return jsonify({
            "valid": False,
            "reason": "not_found",
            "qr_name":  qr_name,
            "qr_class": qr_class,
            "qr_id":    qr_id,
        })

    # Cross-check name too (soft warning, not a hard reject)
    name_ok = match["name"].lower() == qr_name.lower()

    return jsonify({
        "valid":    True,
        "arrived":  match["arrived"],
        "name_ok":  name_ok,
        "row":      match["row"],
        "ticket": {
            "name":   match["name"],
            "class_": match["class_"],
            "id":     match["id"],
            "email":  match["email"],
        },
        "qr": {
            "name":   qr_name,
            "class_": qr_class,
            "id":     qr_id,
        },
    })


@app.route("/api/mark_arrived", methods=["POST"])
def mark_arrived():
    """Body: { "row": <int>, "arrived": true|false }"""
    if not os.path.exists(EXCEL_FILE):
        return jsonify({"error": "Database not found"}), 500

    data = request.get_json(force=True)
    row = data.get("row")
    arrived = data.get("arrived", True)

    if not row:
        return jsonify({"error": "Missing row"}), 400

    if arrived:
        _mark_arrived(row)
    else:
        _unmark_arrived(row)

    return jsonify({"ok": True, "row": row, "arrived": arrived})


@app.route("/api/stats")
def stats():
    """Quick dashboard numbers."""
    if not os.path.exists(EXCEL_FILE):
        return jsonify({"error": "Database not found"}), 500

    _ensure_arrived_header()
    students = _load_all()
    total    = len(students)
    arrived  = sum(1 for s in students if s["arrived"])
    return jsonify({
        "total":   total,
        "arrived": arrived,
        "pending": total - arrived,
    })


@app.route("/api/attendees")
def attendees():
    """Full list for the dashboard."""
    if not os.path.exists(EXCEL_FILE):
        return jsonify({"error": "Database not found"}), 500

    _ensure_arrived_header()
    students = _load_all()
    # Remove internal row key from public API response
    for s in students:
        pass  # keep row for mark_arrived calls
    return jsonify(students)


if __name__ == "__main__":
    print("=" * 55)
    print("  🎟  Ticket Verifier")
    print("=" * 55)
    if not os.path.exists(EXCEL_FILE):
        print(f"  ⚠  WARNING: '{EXCEL_FILE}' not found in current directory.")
        print(f"     Copy your spreadsheet here before scanning tickets.")
    else:
        print(f"  ✓  Database: {EXCEL_FILE}")
    print()
    print("  Open in browser (same WiFi):")
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        print(f"  → https://{ip}:5000")
        print()
        qr = qrcode.QRCode(border=1)
        qr.add_data(f"https://{ip}:5000")
        qr.make(fit=True)
        qr.print_ascii(invert=True)
    except Exception:
        print("  → http://localhost:5000")
    print("=" * 55)
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True, ssl_context=("cert.pem", "key.pem"))